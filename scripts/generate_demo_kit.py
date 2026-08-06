"""Generate Tempolith screen-recording demo workbooks and production plan.

Run from the repository root with the bundled or project Python runtime.
All data is deterministic and synthetic.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import shutil
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "tempolith-demo-kit"
DATA_DIR = OUT / "datasets"
RNG = np.random.default_rng(20260714)


@dataclass(frozen=True)
class Demo:
    number: int
    title: str
    message: str
    duration: str
    status: str
    workbook: str
    mapping: str
    setup: str
    screen_flow: str
    proof: str
    closing: str


DEMOS = [
    Demo(1, "Excel ingestion and column mapping", "Bring an ordinary business workbook into Tempolith in under a minute.", "45-60 sec", "Record now", "01_excel_ingestion_and_mapping.xlsx", "Sheet: Monthly_Data; Date: Date; Value: Revenue; Series: none", "Start on Ingest with no active dataset.", "Upload workbook; choose Monthly_Data; show preview; confirm automatic mapping; open the saved dataset from the library.", "The workbook is previewed, mapped, stored locally, and reusable without another upload.", "Your forecasting workflow starts with the spreadsheet you already have."),
    Demo(2, "Forecast readiness report", "Check whether a series is forecast-ready before trusting a model.", "50-70 sec", "Record now", "02_forecast_readiness.xlsx", "Date: Date; Value: Daily_Demand; Series: none", "Upload, then open Data Quality.", "Run preflight; show quality score; show frequency, trend, seasonality, outliers, stationarity, and recommended transformations.", "The dataset deliberately contains weekly seasonality, trend, and more than 5% extreme points.", "A forecast is only as defensible as the data behind it."),
    Demo(3, "Forecast horizon and uncertainty", "Get a forward forecast with a realistic range, not a single false-precision line.", "55-75 sec", "Record now", "03_forecast_horizon_and_uncertainty.xlsx", "Date: Date; Value: Revenue; Series: none", "Upload and open Forecast; wait for model status Ready.", "Choose horizon 12; run forecast; hover history and forecast; point to P10/P90 band, expected total, accuracy, and confidence.", "A seasonal monthly series produces a clear 12-month forecast and uncertainty band.", "Plan with a range of outcomes, not one number."),
    Demo(4, "TimesFM versus LightGBM", "Let backtested evidence choose between two different forecasting approaches.", "60-80 sec", "Record now", "04_model_comparison.xlsx", "Date: Date; Value: Orders; Series: none", "Upload and open Forecast.", "Run horizon 28; show recommended winner; inspect the alternative; toggle side-by-side overlay; mention holdout MAPE and confidence.", "Both forecasts, the selected winner, and the alternative are visible on the same history.", "Tempolith does not ask you to choose a model by instinct."),
    Demo(5, "Forecast export", "Turn an analysis into an asset you can share immediately.", "35-50 sec", "Record now", "05_forecast_export.xlsx", "Date: Date; Value: Revenue; Series: none", "Run a 12-period forecast first.", "Export the chart as PNG; download the forecast PDF; briefly open the downloaded file if your recording setup allows it.", "The chart export and structured PDF complete successfully.", "The result leaves Tempolith in a format stakeholders can use."),
    Demo(6, "Walk-forward backtesting", "Prove how the forecast would have performed on history it had not seen.", "70-90 sec", "Record now", "06_walk_forward_backtest.xlsx", "Date: Date; Value: Daily_Orders; Series: none", "Upload and open Backtest; wait for model Ready.", "Set horizon 14 and folds 5; run; show winner, aggregate MAPE/RMSE/MASE, per-horizon error, and fold table.", "Stable weekly demand gives readable fold metrics and a clear error-by-horizon chart.", "A plausible forecast is not enough. Backtest it."),
    Demo(7, "Prediction interval calibration", "Check whether the forecast range is as honest as the point forecast.", "55-75 sec", "Record now", "07_prediction_interval_calibration.xlsx", "Date: Date; Value: Shipments; Series: none", "Run Backtest with horizon 14 and folds 5.", "Click Compute calibration; explain nominal versus empirical coverage; point out whether intervals are under or overconfident.", "The reliability plot and calibration observations render after the backtest.", "Uncertainty should be measured, not decorated onto a chart."),
    Demo(8, "Residual diagnostics", "See what the model failed to learn.", "70-90 sec", "Record now", "08_residual_diagnostics.xlsx", "Date: Date; Value: Support_Calls; Series: none", "Upload and open Diagnostics.", "Choose horizon 14 and Seasonal Naive; run; show histogram, Q-Q plot, residual ACF, Ljung-Box result, and STL decomposition.", "The dataset has trend, weekly structure, and periodic shocks, so the diagnostic panels are visually informative.", "Good diagnostics tell you where confidence should stop."),
    Demo(9, "Contextual anomaly detection", "Find the dates that were genuinely unusual relative to the underlying pattern.", "55-75 sec", "Record now", "09_contextual_anomaly_detection.xlsx", "Date: Date; Value: Transactions; Series: none", "Upload and open Anomalies.", "Choose look-ahead 14; run; show critical and warning counts, pulsing markers, biggest spike, biggest drop, and flagged table.", "Known synthetic spikes and drops should appear as visible critical points.", "Start investigations with the few dates that actually deserve attention."),
    Demo(10, "Multi-method anomaly agreement", "Use method agreement to separate robust anomalies from one-method noise.", "60-80 sec", "Fix first", "10_multi_method_anomaly_agreement.xlsx", "Date: Date; Value: Sales; Series: none", "Fix quantile-PI wiring before recording, then open Explain.", "Run Detect anomalies; show method counts, agreement matrix, vote severity, and reasons.", "Extreme campaign days are designed to be found by several statistical methods.", "Agreement across methods is stronger evidence than one arbitrary threshold."),
    Demo(11, "Lag analysis", "Measure how many periods pass before a business driver shows up in the target.", "50-70 sec", "Record now", "11_lag_analysis.xlsx", "Date: Date; Value: Sales; numeric factors: Ad_Spend, Price, Stockout_Flag, Inventory_Index", "Upload and open Explain; map columns; select Ad_Spend and Price.", "Click Lag analysis; focus on the Ad_Spend curve and peak lag; explain that positive lag means the factor leads Sales.", "Sales is generated with a three-day delayed response to Ad_Spend.", "Knowing the lag turns a correlation into an operating lead time."),
    Demo(12, "Granger causality", "Test whether a driver predicts the target beyond the target's own history.", "50-70 sec", "Record now", "12_granger_causality.xlsx", "Date: Date; Value: Sales; numeric factors: Ad_Spend, Price, Stockout_Flag, Inventory_Index", "Upload and open Explain; select Ad_Spend and Price.", "Click Granger causality; show best lag, p-value, and causal flag; state that this is predictive evidence, not philosophical proof of cause.", "The delayed Ad_Spend signal is intentionally strong enough to be detected.", "Use causality tests to narrow the drivers worth acting on."),
    Demo(13, "Anomaly root-cause hints", "Connect abnormal target dates to the factors that were abnormal at the same time.", "65-85 sec", "Fix first", "13_anomaly_root_cause.xlsx", "Date: Date; Value: Sales; numeric: Ad_Spend, Stockout_Flag, Inventory_Index; categorical: Event_Type", "Fix quantile-PI wiring; open Explain; select factors.", "Run anomaly methods; then Find root cause; show anomaly versus baseline means, factor z-scores, category lift, direction, and strength.", "Anomalies coincide with stockouts, unusually low inventory, and distinct event categories.", "Tempolith gives you evidence-backed investigation leads, not an invented narrative."),
    Demo(14, "Factor impact", "Quantify how price, spend, promotions, and weather change the forecast.", "70-90 sec", "Record now", "14_factor_impact.xlsx", "Date: Date; Value: Sales; numeric: Ad_Spend, Price, Temperature_C; categorical: Promotion", "Upload and open Factors.", "Select factors; choose additive; run horizon 12; show baseline toggle, total delta, top driver, influence bars, correlations, and elasticity table.", "The target contains clear positive spend and promotion effects plus a negative price effect.", "A factor belongs in planning only when its impact is visible and measurable."),
    Demo(15, "What-if scenarios", "Compare a flat future with a spend ramp and a price change before committing resources.", "75-95 sec", "Record now", "15_what_if_scenarios.xlsx", "Date: Date; Value: Sales; numeric factors: Ad_Spend, Price", "Upload and open What-if.", "Select Ad_Spend and Price; run baseline; add a 12-period spend ramp; add a price scenario; compare totals and percentage deltas.", "The scenario chart shows multiple futures against the same historical series.", "A scenario is useful when it makes a decision comparable."),
    Demo(16, "Segment comparison", "Find which regions are largest, fastest-growing, and most volatile.", "55-75 sec", "Record now", "16_segment_comparison.xlsx", "Date: Date; Value: Demand; Series: Segment", "Upload and open Segments; select Segment as series column.", "Set Top N 10; compare; switch sort between Total, Growth, and Volatility; show multi-line timelines.", "Eight segments have intentionally different size, growth, seasonality, and volatility profiles.", "The biggest segment is not always the one that needs the most attention."),
    Demo(17, "Timeline annotations", "Put known business events next to the forecasting evidence.", "40-60 sec", "Record now", "17_timeline_annotations.xlsx", "Date: Date; Value: Revenue; Series: none", "Upload and open Operations.", "Add annotations for 2024-03-01 Product launch and 2024-11-01 Distribution expansion; show saved entries; delete one and add it again if useful.", "Annotations persist against the active dataset.", "Known events should be recorded before unusual movements are interpreted."),
    Demo(18, "Saved analyses and operations PDF", "Keep a local audit trail of the analyses behind a decision.", "55-75 sec", "Record now", "18_saved_analyses_and_ops_pdf.xlsx", "Date: Date; Value: Revenue; Series: none", "Run Preflight and Backtest first, then open Operations.", "Show saved analysis entries; export the Operations PDF; explain that annotations and cached analyses are bundled locally.", "Preflight and backtest runs appear automatically in Saved analyses.", "The decision record stays with the dataset on your machine."),
    Demo(19, "SQL connection and table ingest", "Connect Tempolith directly to a forecasting table without exporting a CSV every time.", "60-90 sec", "Setup required", "19_sql_connection_seed.xlsx", "Database table columns: Date, Demand, Region; Date: Date; Value: Demand; Series: Region", "Import the seed workbook into a local PostgreSQL, MySQL, or SQL Server instance before recording.", "Create connection; test it; choose schema and table; ingest; map Date, Demand, and Region; preview the stored dataset.", "A live local database is required. The workbook is only the seed source for that table.", "Tempolith can start from the system where the time series already lives."),
    Demo(20, "Changepoint detection", "Detect sustained shifts in the level of a series, not just isolated spikes.", "50-70 sec", "Fix first", "20_changepoint_detection.xlsx", "Date: Date; Value: Demand; Series: none", "Add ruptures to the app dependency set and verify installation before recording.", "Open Explain; run Detect changepoints; show shift date, left versus right mean, direction, and percent change.", "The series contains two deliberate level shifts.", "A changepoint tells you when the old baseline stopped being valid."),
    Demo(21, "Missing-value quality guardrail", "Surface incomplete target data before any forecast is trusted.", "40-60 sec", "Fix first", "21_missing_value_quality_guardrail.xlsx", "Date: Date; Value: Demand; Series: none", "Change preflight so missing values are counted before extraction rejects them, or return a dedicated quality result.", "Run Data Quality; show missing count, missing rate, score penalty, and corrective warning.", "The workbook contains 15 missing target values across 180 daily rows.", "Missing periods should be a visible quality decision, not an opaque failure."),
]


def base_monthly(n: int = 120, start: str = "2016-01-01") -> pd.DataFrame:
    dates = pd.date_range(start, periods=n, freq="MS")
    i = np.arange(n)
    seasonal = 1 + 0.16 * np.sin(2 * np.pi * (dates.month.to_numpy() - 3) / 12)
    revenue = (180_000 + 2_900 * i) * seasonal + RNG.normal(0, 8_000, n)
    return pd.DataFrame({"Date": dates, "Revenue": np.round(revenue, 2)})


def forecast_readiness() -> pd.DataFrame:
    n = 420
    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    i = np.arange(n)
    weekly = 450 * np.sin(2 * np.pi * i / 7)
    trend = 2.2 * i
    values = 1_250 + trend + weekly + RNG.normal(0, 25, n)
    outlier_idx = np.arange(14, n, 18)
    values[outlier_idx] += np.where((np.arange(len(outlier_idx)) % 2) == 0, 2_500, -2_200)
    return pd.DataFrame({"Date": dates, "Daily_Demand": np.round(np.clip(values, 50, None), 2)})


def daily_orders(n: int, start: str, value_name: str, base: float = 1800) -> pd.DataFrame:
    dates = pd.date_range(start, periods=n, freq="D")
    i = np.arange(n)
    dow = dates.dayofweek.to_numpy()
    weekly = np.where(dow >= 5, 0.72, 1.08)
    yearly = 1 + 0.10 * np.sin(2 * np.pi * i / 365)
    trend = 1 + 0.00045 * i
    values = base * weekly * yearly * trend + RNG.normal(0, base * 0.045, n)
    return pd.DataFrame({"Date": dates, value_name: np.round(np.clip(values, 1, None), 2)})


def calibration_data() -> pd.DataFrame:
    n = 900
    df = daily_orders(n, "2023-01-01", "Shipments", 2400)
    scale = np.linspace(30, 180, n)
    df["Shipments"] = np.round(np.clip(df["Shipments"].to_numpy() + RNG.normal(0, scale), 1, None), 2)
    return df


def diagnostics_data() -> pd.DataFrame:
    n = 730
    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    i = np.arange(n)
    weekly = 240 * np.sin(2 * np.pi * (i + 1) / 7)
    biweekly = 85 * np.sin(2 * np.pi * i / 14)
    trend = 900 + 0.55 * i
    calls = trend + weekly + biweekly + RNG.normal(0, 70, n)
    calls[np.arange(60, n, 91)] += 420
    return pd.DataFrame({"Date": dates, "Support_Calls": np.round(np.clip(calls, 20, None), 2)})


def contextual_anomalies() -> pd.DataFrame:
    df = daily_orders(730, "2024-01-01", "Transactions", 3200)
    values = df["Transactions"].to_numpy(copy=True)
    for idx, mult in {85: 2.2, 179: 0.25, 265: 2.45, 361: 0.30, 488: 2.1, 612: 0.20}.items():
        values[idx] *= mult
    df["Transactions"] = np.round(values, 2)
    return df


def causal_factors() -> pd.DataFrame:
    n = 520
    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    i = np.arange(n)
    ad = 85 + 12 * np.sin(2 * np.pi * i / 30) + RNG.normal(0, 6, n)
    price = 10.4 + 0.35 * np.sin(2 * np.pi * i / 120) + 0.0012 * i
    stockout = np.zeros(n)
    inventory = 100 + 8 * np.sin(2 * np.pi * i / 45) + RNG.normal(0, 3, n)
    event = np.full(n, "Normal", dtype=object)
    flash_days = [96, 173, 251, 329, 407, 484]
    for d in flash_days:
        ad[d - 3] += 120
        event[d] = "Flash_Campaign"
    sales = np.zeros(n)
    for t in range(n):
        lagged_ad = ad[t - 3] if t >= 3 else ad[0]
        weekly = 75 * np.sin(2 * np.pi * t / 7)
        sales[t] = 1_150 + 5.8 * lagged_ad - 42 * price[t] + weekly + RNG.normal(0, 38)
    for d in flash_days:
        sales[d] += 1_250
    for d in [142, 366]:
        stockout[d] = 1
        inventory[d] = 22
        sales[d] -= 650
        event[d] = "Stockout"
    return pd.DataFrame({
        "Date": dates,
        "Sales": np.round(np.clip(sales, 1, None), 2),
        "Ad_Spend": np.round(ad, 2),
        "Price": np.round(price, 2),
        "Stockout_Flag": stockout.astype(int),
        "Inventory_Index": np.round(inventory, 2),
        "Event_Type": event,
    })


def factor_data() -> pd.DataFrame:
    n = 120
    dates = pd.date_range("2016-01-01", periods=n, freq="MS")
    i = np.arange(n)
    ad = 42_000 + 260 * i + 7_000 * np.sin(2 * np.pi * i / 12) + RNG.normal(0, 2_200, n)
    price = 8.8 + 0.018 * i + 0.12 * np.sin(2 * np.pi * i / 18)
    temp = 24 + 10 * np.sin(2 * np.pi * (i - 3) / 12) + RNG.normal(0, 1.2, n)
    promo_flag = ((i % 6) == 0) | ((i % 12) == 10)
    promo = np.where(promo_flag, "Promotion", "No_Promotion")
    sales = (
        210_000
        + 2.65 * ad
        - 7_500 * price
        + 18_000 * promo_flag.astype(float)
        + 1_100 * temp
        + 1_350 * i
        + RNG.normal(0, 7_500, n)
    )
    return pd.DataFrame({
        "Date": dates,
        "Sales": np.round(sales, 2),
        "Ad_Spend": np.round(ad, 2),
        "Price": np.round(price, 2),
        "Temperature_C": np.round(temp, 2),
        "Promotion": promo,
    })


def segment_data() -> pd.DataFrame:
    dates = pd.date_range("2019-01-01", periods=84, freq="MS")
    i = np.arange(len(dates))
    specs = [
        ("Cairo_Core", 125_000, 1_500, 0.05),
        ("Alexandria", 92_000, 1_050, 0.08),
        ("Delta", 78_000, 1_700, 0.10),
        ("Upper_Egypt", 61_000, 1_900, 0.13),
        ("Canal", 48_000, 750, 0.06),
        ("Red_Sea", 35_000, 1_200, 0.22),
        ("North_Coast", 24_000, 600, 0.38),
        ("Online", 19_000, 2_450, 0.18),
    ]
    frames = []
    for j, (name, base, slope, volatility) in enumerate(specs):
        season = 1 + (0.10 + j * 0.015) * np.sin(2 * np.pi * (i - j) / 12)
        values = (base + slope * i) * season + RNG.normal(0, base * volatility, len(i))
        frames.append(pd.DataFrame({"Date": dates, "Demand": np.round(np.clip(values, 100, None), 2), "Segment": name}))
    return pd.concat(frames, ignore_index=True)


def changepoint_data() -> pd.DataFrame:
    n = 480
    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    level = np.select([np.arange(n) < 150, np.arange(n) < 320], [1_100, 1_520], default=1_240)
    values = level + 95 * np.sin(2 * np.pi * np.arange(n) / 7) + RNG.normal(0, 38, n)
    return pd.DataFrame({"Date": dates, "Demand": np.round(values, 2)})


def missing_data() -> pd.DataFrame:
    df = daily_orders(180, "2025-01-01", "Demand", 1450)
    miss = [8, 19, 27, 41, 58, 66, 79, 91, 104, 119, 132, 145, 157, 168, 176]
    df.loc[miss, "Demand"] = np.nan
    return df


def annotation_data() -> pd.DataFrame:
    df = base_monthly(84, "2019-01-01")
    df["Known_Event"] = ""
    df.loc[df["Date"] == pd.Timestamp("2024-03-01"), "Known_Event"] = "Product launch"
    df.loc[df["Date"] == pd.Timestamp("2024-11-01"), "Known_Event"] = "Distribution expansion"
    return df


def sql_seed() -> pd.DataFrame:
    frames = []
    dates = pd.date_range("2025-01-01", periods=240, freq="D")
    for region, base, growth in [("North", 900, 0.5), ("South", 700, 0.8), ("East", 560, 1.0)]:
        i = np.arange(len(dates))
        values = base + growth * i + 90 * np.sin(2 * np.pi * i / 7) + RNG.normal(0, 28, len(i))
        frames.append(pd.DataFrame({"Date": dates, "Demand": np.round(values, 2), "Region": region}))
    return pd.concat(frames, ignore_index=True)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="112127")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for idx, column in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 28)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, pd.Timestamp):
                    cell.number_format = "yyyy-mm-dd"
                elif hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
    wb.save(path)


def write_workbook(filename: str, df: pd.DataFrame, dictionary: pd.DataFrame | None = None) -> None:
    path = DATA_DIR / filename
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        sheet = "Monthly_Data" if filename.startswith("01_") else "Data"
        df.to_excel(writer, sheet_name=sheet, index=False)
        if dictionary is not None:
            dictionary.to_excel(writer, sheet_name="Data_Dictionary", index=False)
    style_workbook(path)


def build_plan_dataframe() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "No": d.number,
            "Title": d.title,
            "Core message": d.message,
            "Duration": d.duration,
            "Status": d.status,
            "Workbook": d.workbook,
            "Mapping": d.mapping,
            "Setup": d.setup,
            "Screen flow": d.screen_flow,
            "On-screen proof": d.proof,
            "Closing line": d.closing,
        }
        for d in DEMOS
    ])


def write_plan_files() -> None:
    plan_df = build_plan_dataframe()
    blockers = pd.DataFrame([
        {"Area": "Preflight missing values", "Finding": "extract_series rejects missing targets before run_preflight can count them", "Recording decision": "Hold demo 21 until fixed"},
        {"Area": "Five-method anomaly agreement", "Finding": "quantile_pi is always empty because existing PI flags are not supplied", "Recording decision": "Hold demos 10 and 13 until fixed"},
        {"Area": "Changepoints", "Finding": "service imports ruptures, but ruptures is absent from project dependencies", "Recording decision": "Hold demo 20 until fixed"},
        {"Area": "SQL connector", "Finding": "requires a running PostgreSQL, MySQL, or SQL Server instance", "Recording decision": "Record only after local test database setup"},
    ])
    plan_path = OUT / "Tempolith_Demo_Production_Plan.xlsx"
    with pd.ExcelWriter(plan_path, engine="openpyxl") as writer:
        plan_df.to_excel(writer, sheet_name="Demo_Plan", index=False)
        blockers.to_excel(writer, sheet_name="Known_Blockers", index=False)
    style_workbook(plan_path)

    lines = [
        "# Tempolith Demo Production Plan",
        "",
        "This plan is based on the implemented React pages, API wiring, and backend services reviewed on 2026-07-14.",
        "",
        "## Recording standard",
        "",
        "- Use one business question per video.",
        "- Target 45 to 90 seconds. Stop when the proof is visible.",
        "- Record at 1920x1080 with browser zoom at 100% and the sidebar open.",
        "- Start with the dataset already uploaded unless ingestion is the feature.",
        "- Wait for the model status to show Ready before starting the take.",
        "- Keep the cursor still while charts animate, then hover only the point being explained.",
        "- End on the result, not on navigation or setup.",
        "",
        "## Series order",
        "",
    ]
    for d in DEMOS:
        lines.extend([
            f"### {d.number:02d}. {d.title}",
            "",
            f"Status: {d.status}",
            "",
            f"Message: {d.message}",
            "",
            f"Workbook: `datasets/{d.workbook}`",
            "",
            f"Mapping: {d.mapping}",
            "",
            f"Setup: {d.setup}",
            "",
            f"Screen flow: {d.screen_flow}",
            "",
            f"Proof to hold on screen: {d.proof}",
            "",
            f"Closing line: {d.closing}",
            "",
        ])
    (OUT / "TEMPOLITH_DEMO_PLAN.md").write_text("\n".join(lines), encoding="utf-8")


def generate() -> None:
    if OUT.exists():
        shutil.rmtree(OUT)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    monthly = base_monthly()
    dictionary = pd.DataFrame({
        "Column": ["Date", "Revenue", "Marketing_Spend", "Promotion"],
        "Meaning": ["Month start", "Monthly net revenue", "Monthly marketing investment", "Promotion status"],
        "Type": ["Date", "Currency", "Currency", "Category"],
    })
    ingest = monthly.iloc[-96:].copy()
    ingest["Marketing_Spend"] = np.round(24_000 + np.arange(96) * 180 + RNG.normal(0, 1_400, 96), 2)
    ingest["Promotion"] = np.where(np.arange(96) % 6 == 0, "Promotion", "None")

    causal = causal_factors()
    factors = factor_data()
    annotation = annotation_data()

    write_workbook(DEMOS[0].workbook, ingest, dictionary)
    write_workbook(DEMOS[1].workbook, forecast_readiness())
    write_workbook(DEMOS[2].workbook, monthly)
    write_workbook(DEMOS[3].workbook, daily_orders(840, "2023-01-01", "Orders", 2100))
    write_workbook(DEMOS[4].workbook, monthly.copy())
    write_workbook(DEMOS[5].workbook, daily_orders(900, "2023-01-01", "Daily_Orders", 1900))
    write_workbook(DEMOS[6].workbook, calibration_data())
    write_workbook(DEMOS[7].workbook, diagnostics_data())
    write_workbook(DEMOS[8].workbook, contextual_anomalies())
    write_workbook(DEMOS[9].workbook, causal.copy())
    write_workbook(DEMOS[10].workbook, causal.copy())
    write_workbook(DEMOS[11].workbook, causal.copy())
    write_workbook(DEMOS[12].workbook, causal.copy())
    write_workbook(DEMOS[13].workbook, factors.copy())
    write_workbook(DEMOS[14].workbook, factors.copy())
    write_workbook(DEMOS[15].workbook, segment_data())
    write_workbook(DEMOS[16].workbook, annotation.copy())
    write_workbook(DEMOS[17].workbook, annotation.copy())
    write_workbook(DEMOS[18].workbook, sql_seed())
    write_workbook(DEMOS[19].workbook, changepoint_data())
    write_workbook(DEMOS[20].workbook, missing_data())

    write_plan_files()

    zip_path = ROOT / "output" / "Tempolith_Demo_Kit.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(OUT.rglob("*")):
            if path.is_file():
                archive.write(path, path.relative_to(OUT.parent))

    print(f"Generated {len(DEMOS)} demo workbooks in {DATA_DIR}")
    print(f"Plan: {OUT / 'TEMPOLITH_DEMO_PLAN.md'}")
    print(f"Archive: {zip_path}")


if __name__ == "__main__":
    generate()





